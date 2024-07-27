import re
from datetime import datetime
from time import sleep

import undetected_chromedriver as uc
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common import StaleElementReferenceException, NoSuchElementException, TimeoutException, \
    ElementClickInterceptedException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait


def setup_driver():
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    # options.add_argument(
    #     "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36")
    # options.add_argument("--incognito")
    options.add_argument("--disable-extensions")
    options.add_argument("--disable-popup-blocking")
    # options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    # options.add_argument("--disable-renderer-backgrounding")
    # options.add_argument("--disable-background-timer-throttling")
    # options.add_argument("--disable-backgrounding-occluded-windows")
    # options.add_argument("--disable-client-side-phishing-detection")
    # options.add_argument("--disable-crash-reporter")
    # options.add_argument("--disable-oopr-debug-crash-dump")
    # options.add_argument("--no-crash-upload")
    # options.add_argument("--disable-gpu")
    # options.add_argument("--disable-low-res-tiling")
    # options.add_argument("--log-level-3") a
    # options.add_argument("--silent")
    options.add_argument("--page-load-strategy=none")
    options.add_argument("--blink-settings=imagesEnabled=false")
    options.add_argument("--disable-features=Images")

    driver = uc.Chrome(options=options)

    driver.maximize_window()
    print("Starting script...")
    time_now = datetime.now()
    formatted_time = time_now.strftime("%Y-%m-%d %H:%M:%S")
    print("----------------Time now:", formatted_time)
    return driver


def is_valid_year_range(series_name):
    # Regex za prepoznavanje formata godina (xxxx - xxxx) i (xxxx - ...)
    match = re.search(r'\((\d{4}) - (\d{4}|\.\.\.)\)', series_name)
    if match:
        start_year = int(match.group(1))
        # Provera da li je početna godina najmanje 1995
        if start_year >= 1995:
            return True
    return False


def initialize_excel(file_path):
    workbook = Workbook()
    sheet = workbook.active
    headers = ["FILTER NAME", "FILTER NUMBER", "FILTER BRAND", "FILTER TYPE", "HEIGHT (mm)", "DIAMETER (mm)",
               "INNER DIAMETER (mm)", "INNER DIAMETER 2 (mm)", "INLET Ø (mm)", "OUTLET Ø (mm)", "PIPE Ø (mm)",
               "CONSTRUCTION YEAR TO", "CONSTRUCTION YEAR FROM", "ENGINE CODE", "ENGINE NUMBER TO",
               "ENGINE NUMBER FROM", "VEHICLE IDENTIFICATION NUMBER (VIN) FROM",
               "VEHICLE IDENTIFICATION NUMBER (VIN) TO", "VEHICLE IDENTIFICATION NUMBER (VIN)",
               "SUPPLEMENTARY ARTICLE / SUPPLEMENTARY INFO", "SUPPLEMENTARY ARTICLE / SUPPLEMENTARY INFO 2",
               "MANUFACTURER RESTRICTION", "HOUSING TYPE", "VEHICLE EQUIPMENT", "EMISSION STANDARD",
               "LEFT/RIGHT HAND DRIVE VEHICLES", "STATUS", "FOR OE NUMBER",
               "CAR BRAND", "CAR MODEL", "CAR SERIES AND YEAR", "CAR ENGINE"]
    sheet.append(headers)
    workbook.save(file_path)


def append_to_excel(file_path, data):
    global sheet
    sheet.append(data)


def close_excel(file_path):
    global workbook
    workbook.save(file_path)
    workbook.close()


def adjust_column_widths(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
    workbook.save(file_path)


def wait_for_url_change(driver, current_url):
    try:
        WebDriverWait(driver, 60).until(
            lambda driver: driver.current_url != current_url
        )
    except TimeoutException:
        try:
            print("TimeoutException for current_url caught.")
            print("Refreshing the page")
            driver.refresh()
            WebDriverWait(driver, 30).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//div[@class="container-fluid"]'))
            )
            print("Page refreshed")
            main_div = driver.find_element(By.XPATH,
                                           './/div[contains(@class, "header-select__choosse-wrap")]')
            search_button = main_div.find_element(By.XPATH, './/button[@type="button" and contains(text(), "Search")]')
            search_button.click()
        except StaleElementReferenceException:
            print("StaleElementReferenceException in wait_for_url_change. Continuing.")


def wait_for_listing_div(driver):
    try:
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, '//div[@class="title-car title-car--page title-car--recomended-block"]'))
        )
    except TimeoutException:
        print("wait_for_listing_div TimeoutException caught.")
        driver.refresh()
        print("Refreshing the page")
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, '//div[@class="container-fluid"]'))
        )
        print("Page refreshed. Going to oil filters.")
        driver.get(f'https://www.onlinecarparts.co.uk/spare-parts/oil-filter.html/')
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, './/div[contains(@class, "header-select__choosse-wrap")]'))
        )


def accept_cookies(driver):
    try:
        # Sačekajte da dugme za prihvatanje kolačića postane vidljivo i kliknite na njega
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//button[@data-cookies="allow_all_cookies"]'))
        ).click()
        print("Cookies accepted.")
    except TimeoutException:
        print("TimeoutException: No cookies acceptance button found.")
    except ElementClickInterceptedException:
        print("Element click intercepted when accepting cookies.")
    except Exception as e:
        print(f"An error occurred while accepting cookies: {e}")


def get_filtered_url(current_url):
    # Uklanjanje "#" i bilo čega posle nje
    if '#' in current_url:
        current_url = current_url.split('#')[0]

    # Dodavanje filtera na URL
    filter_query = "?brand%5B%5D=254&brand%5B%5D=30&brand%5B%5D=4&brand%5B%5D=38&brand%5B%5D=81&brand%5B%5D=10565"
    filtered_url = current_url + filter_query

    return filtered_url


def online_car_parts(driver, file_path):
    # Otvaranje stranice
    driver.get(f'https://www.onlinecarparts.co.uk/spare-parts/fuel-filter.html/')
    accept_cookies(driver)

    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, './/div[contains(@class, "header-select__choosse-wrap")]'))
    )

    # Pronalaženje glavnog div-a
    main_div = driver.find_element(By.XPATH, './/div[contains(@class, "header-select__choosse-wrap")]')

    # Pronalaženje svih div-ova unutar main_div-a koji imaju "selector" u klasi
    selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')

    # Rad sa prvim div-om koji ima klasu "selector"
    first_selector_div = selector_divs[0]
    third_selector_div = selector_divs[2]

    # Pronalaženje select elementa unutar prvog div-a
    select_element_brand = first_selector_div.find_element(By.TAG_NAME, 'select')

    # Pronalaženje optgroup elementa sa labelom "Carmakers are arranged in alphabetical order"
    alphabetical_optgroup_brand = select_element_brand.find_element(By.XPATH,
                                                                    './/optgroup[@label="Carmakers are arranged in alphabetical order"]')

    # Lista brendova koji vas zanimaju
    # desired_brands = {
    #     "ABARTH", "ALFA ROMEO", "AUDI", "BMW", "CHEVROLET", "CITROЁN", "CUPRA", "DACIA",
    #     "DAEWOO", "DS", "FIAT", "FORD", "HONDA", "HYUNDAI", "INFINITI", "IVECO", "JAGUAR",
    #     "JEEP", "KIA", "LADA", "LAMBORGHINI", "LANCIA", "LAND ROVER", "LEXUS", "MAN",
    #     "MAZDA", "MERCEDES-BENZ", "MINI", "MITSUBISHI", "NISSAN", "OPEL", "PEUGEOT",
    #     "PORSCHE", "RENAULT", "ROVER", "SAAB", "SEAT", "SKODA", "SMART", "SUBARU", "SUZUKI",
    #     "TESLA", "TOYOTA", "VW", "VOLVO"
    # }

    desired_brands = {
        "ABARTH"
    }

    # Pronalaženje svih opcija unutar alphabetical_optgroup elementa
    options_brand = alphabetical_optgroup_brand.find_elements(By.TAG_NAME, 'option')

    second_selector_div = selector_divs[1]
    select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
    models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')

    # Ispisivanje tekstova svih opcija koje su u listi željenih brendova
    for i in range(len(options_brand)):
        options_brand = alphabetical_optgroup_brand.find_elements(By.TAG_NAME, 'option')
        option_brand = options_brand[i]
        if option_brand.text in desired_brands:
            select_element_brand.click()
            option_brand.click()
            brand_name = option_brand.text
            sleep(0.5)

            # Ponovno pronalaženje elementa nakon klika
            try:
                main_div = driver.find_element(By.XPATH, './/div[contains(@class, "header-select__choosse-wrap")]')
                selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                second_selector_div = selector_divs[1]
                select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
                models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
            except IndexError as e:
                print(f"[240] IndexError: {e}")
                driver.refresh()
                sleep(5)
                main_div = driver.find_element(By.XPATH,
                                               './/div[contains(@class, "header-select__choosse-wrap")]')
                selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                second_selector_div = selector_divs[1]
                select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
                models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
            for j in range(len(models)):
                models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
                model = models[j]
                model_name = model.get_attribute('label')
                options_series = model.find_elements(By.TAG_NAME, 'option')
                sleep(0.5)
                series_name = None
                try:
                    for k in range(len(options_series)):
                        # options_series = model.find_elements(By.TAG_NAME, 'option')
                        option_series = options_series[k]
                        # Provera validnosti serije pre ekstrakcije imena
                        if is_valid_year_range(option_series.text):
                            series_name = option_series.text
                            option_series.click()
                            sleep(0.5)

                        if series_name:
                            # Ponovno pronalaženje elementa za motor
                            main_div = driver.find_element(By.XPATH,
                                                           './/div[contains(@class, "header-select__choosse-wrap")]')
                            selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                            engine_selector_div = selector_divs[2]
                            select_element_engine = engine_selector_div.find_element(By.TAG_NAME, 'select')
                            options_engine = select_element_engine.find_elements(By.TAG_NAME, 'option')
                            for l in range(len(options_engine)):
                                try:
                                    options_engine = select_element_engine.find_elements(By.TAG_NAME, 'option')
                                    if l < len(options_engine):
                                        option_engine = options_engine[l]
                                        if option_engine.get_attribute('value') != "-1":
                                            engine_name = option_engine.text
                                            print(f"{brand_name} - {model_name} - {series_name} - {engine_name}")

                                            try:
                                                option_engine.click()
                                            except ElementClickInterceptedException:
                                                accept_cookies(driver)
                                                sleep(1)
                                                option_engine.click()
                                            except Exception as e:
                                                print(
                                                    f"An error occurred while clicking option_engine click button: {e}")
                                            # Dodavanje podataka u Excel fajl
                                            # append_to_excel(file_path, [brand_name, model_name, series_name, engine_name])

                                            try:
                                                search_button = main_div.find_element(By.XPATH,
                                                                                      './/button[@type="button" and contains(text(), "Search")]')
                                                search_button.click()
                                            except ElementClickInterceptedException:
                                                accept_cookies(driver)
                                                sleep(1)
                                                search_button = main_div.find_element(By.XPATH,
                                                                                      './/button[@type="button" and contains(text(), "Search")]')
                                                search_button.click()
                                            except Exception as e:
                                                print(f"An error occurred while clicking search button: {e}")

                                            current_url = driver.current_url
                                            wait_for_url_change(driver, current_url)

                                            # Čekanje pojave listing div-a
                                            wait_for_listing_div(driver)

                                            # Sleep 0.5 sekundi nakon što se pojavi listing div
                                            # sleep(0.5)
                                            # WebDriverWait(driver, 30).until(
                                            #     EC.visibility_of_element_located((By.XPATH, '//div[@class="filters-wrapper" and @data-listing-filters=""]'))
                                            # )

                                            WebDriverWait(driver, 60).until(
                                                EC.visibility_of_element_located((By.XPATH,
                                                                                  '//div[@class="col col-md-12 col-xl-10 pl-0 order-2 content-page"]'))
                                            )

                                            # Definisanje brendova za pretragu, uključujući FILTRON
                                            brands_to_search = ["FILTRON", "BOSCH", "MANN-FILTER", "BLUE PRINT",
                                                                "PURFLUX", "HENGST"]

                                            found_brands = []
                                            product_main_divs = []

                                            # Kliknuti next_button odmah na početku
                                            try:
                                                current_url = driver.current_url
                                                filtered_url = get_filtered_url(current_url)

                                                # Učitavanje filtriranog URL-a
                                                driver.get(filtered_url)
                                                # WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                                #     (By.XPATH, '//div[@class="container-fluid"]')))
                                                WebDriverWait(driver, 60).until(EC.invisibility_of_element_located(
                                                    (By.XPATH, '//a[@class="listing-pagination__next-wrap active"]')))
                                            except TimeoutException:
                                                print("TimeoutException. Refreshing the page")
                                                driver.refresh()
                                                WebDriverWait(driver, 30).until(EC.presence_of_element_located(
                                                    (By.XPATH, '//div[@class="container-fluid"]')))
                                            except NoSuchElementException:
                                                print("No more pages.")
                                            except ElementClickInterceptedException:
                                                print("Element click intercepted.")

                                            # Iteracija kroz sve brendove
                                            for brand in brands_to_search:
                                                try:
                                                    # Pronađi sve `div` elemente unutar glavnog kontejnera
                                                    listing_wrapper_divs = driver.find_elements(By.XPATH,
                                                                                                '//div[@class="listing-wrapper"]/div')

                                                    product_cards = []
                                                    # Iteriraj kroz sve pronađene `div` elemente
                                                    for div in listing_wrapper_divs:
                                                        try:
                                                            div_class = div.get_attribute("class")
                                                            # Ako nađemo `title-recommendation` klasu, prekidamo petlju
                                                            if "title-recommendation" in div_class:
                                                                break
                                                            # Ako `div` ima `product-card` klasu, dodaj ga u listu
                                                            if "product-card" in div_class:
                                                                product_cards.append(div)
                                                        except StaleElementReferenceException:
                                                            print(
                                                                "StaleElementReferenceException: couldn't process a div element.")
                                                            continue
                                                except StaleElementReferenceException:
                                                    print(
                                                        "StaleElementReferenceException: product_cards couldn't be found.")
                                                    continue  # Nastavi sa sledećim brendom ako trenutni nije pronađen

                                                for card in product_cards:
                                                    try:
                                                        product_title = card.find_element(By.XPATH,
                                                                                          './/div[@class="product-card__title"]//*[self::a or self::span]').text

                                                        if brand in product_title:
                                                            print(f"* {brand} found")
                                                            found_brands.append(brand)
                                                            product_main_divs.append(card)  # Čuvamo main div u listu
                                                            # break  # Nastavi sa sledećim brendom čim pronađe trenutni
                                                    except NoSuchElementException:
                                                        print(
                                                            f"NoSuchElementException: no {brand} available. Continue.")  # Nastavi sa sledećim elementom ako trenutni nije pronađen
                                                    except StaleElementReferenceException:
                                                        print(
                                                            "StaleElementReferenceException: product_title couldn't be found.")

                                            # Ispis poruke za pronađene brendove
                                            for found_brand, product_main_div in zip(found_brands, product_main_divs):
                                                print(f"************* {found_brand} scraping data ************")
                                                if product_main_div:
                                                    try:

                                                        try:
                                                            article_number_div = product_main_div.find_element(By.XPATH,
                                                                                                               './/div[@class="product-card__artkl"]')
                                                            article_number = article_number_div.find_element(
                                                                By.TAG_NAME,
                                                                'span').text.strip().replace(
                                                                " ", "")
                                                            print(f"- Article №: {article_number}")
                                                        except NoSuchElementException:
                                                            article_number = "Unknown"
                                                            print("- [404] Article number not found.")

                                                        # filter name
                                                        try:
                                                            fuel_filter_name_element = product_main_div.find_element(
                                                                By.XPATH,
                                                                './/div[@class="product-card__title"]//*[self::a[contains(@class, "product-card__title-link")] or self::span[contains(@class, "product-card__title-link")]]')
                                                            fuel_filter_name = \
                                                                fuel_filter_name_element.text.split('\n')[
                                                                    0].strip()
                                                            print(f"- Original Air Filter name: {fuel_filter_name}")

                                                            if article_number in fuel_filter_name:
                                                                fuel_filter_name = fuel_filter_name.replace(
                                                                    article_number,
                                                                    '').strip()

                                                            print(f"- Filter name: {fuel_filter_name}")
                                                        except NoSuchElementException:
                                                            print("- [404] Filter name not found.")
                                                            continue

                                                        # filter brand
                                                        try:
                                                            fuel_filter_brand_name = found_brand
                                                            print(f"- Filter brand: {fuel_filter_brand_name}")
                                                        except NoSuchElementException:
                                                            fuel_filter_brand_name = "Unknown"
                                                            print("- [404] Filter brand not found.")

                                                        ul_element = None

                                                        # filter type
                                                        try:
                                                            desc_table_div = WebDriverWait(product_main_div, 10).until(
                                                                EC.presence_of_element_located((By.XPATH,
                                                                                                './/div[contains(@class, "product-card__desc-table")]'))
                                                            )

                                                            more_button = None
                                                            try:
                                                                more_button = desc_table_div.find_element(By.XPATH,
                                                                                                          './/div[@class="product-desc-more"]')
                                                            except NoSuchElementException:
                                                                print("No 'More+' button located.")

                                                            if more_button:
                                                                print("'More+' button located.")
                                                                driver.execute_script(
                                                                    "arguments[0].scrollIntoView(true); window.scrollBy(0, -300);",
                                                                    more_button)
                                                                more_button.click()
                                                                sleep(0.5)
                                                                desc_table_div = product_main_div.find_element(By.XPATH,
                                                                                                               './/div[contains(@class, "product-card__desc-table")]')

                                                            ul_element = desc_table_div.find_element(By.XPATH, './ul')
                                                            filter_type_li = ul_element.find_element(By.XPATH,
                                                                                                     './li[contains(span[@class="left"], "Filter type")]')
                                                            fuel_filter_type = filter_type_li.find_element(By.XPATH,
                                                                                                           './span[@class="right"]').text
                                                            print(f"- Filter type: {fuel_filter_type}")
                                                        except NoSuchElementException:
                                                            fuel_filter_type = ""
                                                            print("- [404] Filter type not found.")
                                                        except TimeoutException:
                                                            fuel_filter_type = ""
                                                            print("- [404] Filter type not found.")

                                                        # desc_table_div = product_main_div.find_element(By.XPATH,
                                                        #                                                './/div[@class="product-card__desc-table "]')
                                                        # ul_element = desc_table_div.find_element(By.XPATH, './ul')

                                                        if ul_element:

                                                            # Height [mm]
                                                            try:
                                                                height_li = ul_element.find_element(By.XPATH,
                                                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Height [mm]")]]')
                                                                fuel_filter_height_mm = height_li.find_element(By.XPATH,
                                                                                                               './span[contains(@class, "right")]').text
                                                                print(f"- Height [mm]: {fuel_filter_height_mm}")
                                                            except NoSuchElementException:
                                                                fuel_filter_height_mm = ""
                                                                print("- [404] Height [mm] not found.")

                                                            # Diameter (mm)
                                                            try:
                                                                diameter_li = ul_element.find_element(By.XPATH,
                                                                                                      './/li[./span[contains(@class, "left") and starts-with(normalize-space(text()), "Diameter [mm]")]]')
                                                                fuel_filter_diameter_mm = diameter_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(f"- Diameter [mm]: {fuel_filter_diameter_mm}")
                                                            except NoSuchElementException:
                                                                fuel_filter_diameter_mm = ""
                                                                print("- [404] Diameter [mm] not found.")

                                                            # Inner diameter
                                                            try:
                                                                fuel_filter_inner_diameter_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and starts-with(normalize-space(text()), "Inner Diameter [mm]")]]')
                                                                fuel_filter_inner_diameter = fuel_filter_inner_diameter_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Inner diameter: {fuel_filter_inner_diameter}")
                                                            except NoSuchElementException:
                                                                fuel_filter_inner_diameter = ""
                                                                print("- [404] Inner diameter not found.")

                                                            # Inner diameter 2
                                                            try:
                                                                fuel_filter_inner_diameter2_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and starts-with(normalize-space(text()), "Inner Diameter 2 [mm]")]]')
                                                                fuel_filter_inner_diameter2 = fuel_filter_inner_diameter2_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Inner diameter 2: {fuel_filter_inner_diameter2}")
                                                            except NoSuchElementException:
                                                                fuel_filter_inner_diameter2 = ""
                                                                print("- [404] Inner diameter 2 not found.")

                                                            # Inlet Ø [mm]
                                                            try:
                                                                inlet_li = ul_element.find_element(By.XPATH,
                                                                                                   './/li[./span[contains(@class, "left") and contains(text(), "Inlet Ø [mm]")]]')
                                                                fuel_filter_inlet_mm = inlet_li.find_element(By.XPATH,
                                                                                                             './span[contains(@class, "right")]').text
                                                                print(f"- Inlet Ø [mm]: {fuel_filter_inlet_mm}")
                                                            except NoSuchElementException:
                                                                fuel_filter_inlet_mm = ""
                                                                print("- [404] Inlet Ø [mm] not found.")

                                                            # Outlet Ø [mm]
                                                            try:
                                                                outlet_li = ul_element.find_element(By.XPATH,
                                                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Outlet Ø [mm]")]]')
                                                                fuel_filter_outlet_mm = outlet_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(f"- Outlet Ø [mm]: {fuel_filter_outlet_mm}")
                                                            except NoSuchElementException:
                                                                fuel_filter_outlet_mm = ""
                                                                print("- [404] Outlet Ø [mm] not found.")

                                                            # Pipe Ø [mm]
                                                            try:
                                                                pipe_li = ul_element.find_element(By.XPATH,
                                                                                                  './/li[./span[contains(@class, "left") and contains(text(), "Pipe Ø [mm]")]]')
                                                                fuel_filter_pipe_mm = pipe_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(f"- Pipe Ø [mm]: {fuel_filter_pipe_mm}")
                                                            except NoSuchElementException:
                                                                fuel_filter_pipe_mm = ""
                                                                print("- [404] Pipe Ø [mm] not found.")

                                                            # construction Year to
                                                            try:
                                                                construction_year_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Construction Year to")]]')
                                                                fuel_filter_construction_year_to = construction_year_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Construction Year to: {fuel_filter_construction_year_to}")
                                                            except NoSuchElementException:
                                                                fuel_filter_construction_year_to = ""
                                                                print("- [404] Construction Year to not found.")

                                                            # construction Year from
                                                            try:
                                                                construction_year_from_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Construction Year from")]]')
                                                                fuel_filter_construction_year_from = construction_year_from_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Construction Year from: {fuel_filter_construction_year_from}")
                                                            except NoSuchElementException:
                                                                fuel_filter_construction_year_from = ""
                                                                print("- [404] Construction Year from not found.")

                                                            # engine code
                                                            try:
                                                                engine_code_li = ul_element.find_element(By.XPATH,
                                                                                                         './/li[./span[contains(@class, "left") and contains(text(), "Engine Code")]]')
                                                                engine_code = engine_code_li.find_element(By.XPATH,
                                                                                                          './span[contains(@class, "right")]').text
                                                                print(f"- Engine Code: {engine_code}")
                                                            except NoSuchElementException:
                                                                engine_code = ""
                                                                print("- [404] Engine Code not found.")

                                                            # engine number to
                                                            try:
                                                                engine_number_to_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Engine Number to")]]')
                                                                fuel_filter_engine_number_to = engine_number_to_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Engine Number to: {fuel_filter_engine_number_to}")
                                                            except NoSuchElementException:
                                                                fuel_filter_engine_number_to = ""
                                                                print("- [404] Engine Number to to not found.")

                                                            # engine number from
                                                            try:
                                                                engine_number_from_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Engine Number from")]]')
                                                                fuel_filter_engine_number_from = engine_number_from_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Engine Number from: {fuel_filter_engine_number_from}")
                                                            except NoSuchElementException:
                                                                fuel_filter_engine_number_from = ""
                                                                print("- [404] Engine Number from not found.")

                                                            # Vehicle Identification Number (VIN) from
                                                            try:
                                                                vin_from_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Vehicle Identification Number (VIN) from:")]]')
                                                                fuel_filter_vin_from = vin_from_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- VIN from: {fuel_filter_vin_from}")
                                                            except NoSuchElementException:
                                                                fuel_filter_vin_from = ""
                                                                print(
                                                                    "- [404] VIN from not found.")

                                                            # Vehicle Identification Number (VIN) to
                                                            try:
                                                                vin_to_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Vehicle Identification Number (VIN) to:")]]')
                                                                fuel_filter_vin_to = vin_to_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- VIN to: {fuel_filter_vin_to}")
                                                            except NoSuchElementException:
                                                                fuel_filter_vin_to = ""
                                                                print(
                                                                    "- [404] VIN to not found.")

                                                            # Vehicle Identification Number (VIN)
                                                            try:
                                                                vin_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Vehicle Identification Number (VIN):")]]')
                                                                fuel_filter_vin = vin_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- VIN: {fuel_filter_vin}")
                                                            except NoSuchElementException:
                                                                fuel_filter_vin = ""
                                                                print(
                                                                    "- [404] VIN not found.")

                                                            # Supplementary Article / Supplementary Info
                                                            try:
                                                                supplementary_article_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Supplementary Article / Supplementary Info:")]]')
                                                                fuel_filter_supplementary_article = supplementary_article_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Supplementary Article/Info 1: {fuel_filter_supplementary_article}")
                                                            except NoSuchElementException:
                                                                fuel_filter_supplementary_article = ""
                                                                print("- [404] Supplementary Article/Info 1 not found.")

                                                            # Supplementary Article / Supplementary Info Info 2
                                                            try:
                                                                supplementary_article2_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Supplementary Article / Supplementary Info Info 2:")]]')
                                                                fuel_filter_supplementary_article2 = supplementary_article2_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Supplementary Article/Info 2: {fuel_filter_supplementary_article2}")
                                                            except NoSuchElementException:
                                                                fuel_filter_supplementary_article2 = ""
                                                                print("- [404] Supplementary Article/Info 2 not found.")

                                                            # Manufacturer Restriction
                                                            try:
                                                                manufacturer_restriction_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Manufacturer Restriction")]]')
                                                                fuel_filter_manufacturer_restriction = manufacturer_restriction_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Manufacturer Restriction: {fuel_filter_manufacturer_restriction}")
                                                            except NoSuchElementException:
                                                                fuel_filter_manufacturer_restriction = ""
                                                                print(
                                                                    "- [404] Manufacturer Restriction not found.")

                                                            # Housing Type
                                                            try:
                                                                housing_type_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Housing Type")]]')
                                                                fuel_filter_housing_type = housing_type_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Housing Type: {fuel_filter_housing_type}")
                                                            except NoSuchElementException:
                                                                fuel_filter_housing_type = ""
                                                                print("- [404] Housing Type not found.")

                                                            # Vehicle Equipment
                                                            try:
                                                                vehicle_equipment_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Vehicle Equipment")]]')
                                                                fuel_filter_vehicle_equipment = vehicle_equipment_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Vehicle Equipment: {fuel_filter_vehicle_equipment}")
                                                            except NoSuchElementException:
                                                                fuel_filter_vehicle_equipment = ""
                                                                print("- [404] Vehicle Equipment not found.")

                                                            # Emission standard
                                                            try:
                                                                emission_standard_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Emission Standard")]]')
                                                                fuel_filter_emission_standard = emission_standard_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Emission Standard: {fuel_filter_emission_standard}")
                                                            except NoSuchElementException:
                                                                fuel_filter_emission_standard = ""
                                                                print("- [404] Emission Standard not found.")

                                                            # Left-/right-hand drive vehicles
                                                            try:
                                                                left_right_hand_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "Left-/right-hand drive vehicles")]]')
                                                                fuel_filter_left_right_hand = left_right_hand_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- Left-/right-hand drive vehicles: {fuel_filter_left_right_hand}")
                                                            except NoSuchElementException:
                                                                fuel_filter_left_right_hand = ""
                                                                print("- [404] Left-/right-hand drive vehicles not found.")

                                                            # status
                                                            try:
                                                                status_element = product_main_div.find_element(
                                                                    By.XPATH,
                                                                    './/div[contains(@class, "product-card__status")]'
                                                                )
                                                                status_text = status_element.text
                                                                print(f"- Status: {status_text}")
                                                            except NoSuchElementException:
                                                                status_text = ""
                                                                print("- [404] Status element not found.")

                                                            # for OE number
                                                            try:
                                                                oe_number_li = ul_element.find_element(
                                                                    By.XPATH,
                                                                    './/li[./span[contains(@class, "left") and contains(text(), "for OE number:")]]')
                                                                fuel_filter_oe_number = oe_number_li.find_element(
                                                                    By.XPATH,
                                                                    './span[contains(@class, "right")]').text
                                                                print(
                                                                    f"- For OE Number: {fuel_filter_oe_number}")
                                                            except NoSuchElementException:
                                                                fuel_filter_oe_number = ""
                                                                print("- [404] For OE Number not found.")

                                                            append_to_excel(file_path,
                                                                            [fuel_filter_name,
                                                                             article_number,
                                                                             fuel_filter_brand_name,
                                                                             fuel_filter_type,
                                                                             fuel_filter_height_mm,
                                                                             fuel_filter_diameter_mm,
                                                                             fuel_filter_inner_diameter,
                                                                             fuel_filter_inner_diameter2,
                                                                             fuel_filter_inlet_mm,
                                                                             fuel_filter_outlet_mm,
                                                                             fuel_filter_pipe_mm,
                                                                             fuel_filter_construction_year_to,
                                                                             fuel_filter_construction_year_from,
                                                                             engine_code,
                                                                             fuel_filter_engine_number_to,
                                                                             fuel_filter_engine_number_from,
                                                                             fuel_filter_vin_from,
                                                                             fuel_filter_vin_to,
                                                                             fuel_filter_vin,
                                                                             fuel_filter_supplementary_article,
                                                                             fuel_filter_supplementary_article2,
                                                                             fuel_filter_manufacturer_restriction,
                                                                             fuel_filter_housing_type,
                                                                             fuel_filter_vehicle_equipment,
                                                                             fuel_filter_emission_standard,
                                                                             fuel_filter_left_right_hand,
                                                                             status_text,
                                                                             fuel_filter_oe_number,
                                                                             brand_name,
                                                                             model_name,
                                                                             series_name,
                                                                             engine_name])
                                                        else:
                                                            print("ul_element not found.")

                                                    except NoSuchElementException:
                                                        print("No products found in the new listing.")
                                                    except StaleElementReferenceException:
                                                        print("StaleElementReferenceException: skipping this filter.")

                                            # Nastavak izvršavanja
                                            try:
                                                main_div = driver.find_element(By.XPATH,
                                                                               './/div[contains(@class, "header-select__choosse-wrap")]')
                                                selector_divs = main_div.find_elements(By.XPATH,
                                                                                       './/div[contains(@class, "selector")]')
                                                engine_selector_div = selector_divs[2]
                                                select_element_engine = engine_selector_div.find_element(By.TAG_NAME,
                                                                                                         'select')
                                                options_engine = select_element_engine.find_elements(By.TAG_NAME,
                                                                                                     'option')
                                            except StaleElementReferenceException as e:
                                                print(f"StaleElementReferenceException: {e}")
                                                driver.refresh()
                                                sleep(5)
                                                main_div = driver.find_element(By.XPATH,
                                                                               './/div[contains(@class, "header-select__choosse-wrap")]')
                                                selector_divs = main_div.find_elements(By.XPATH,
                                                                                       './/div[contains(@class, "selector")]')
                                                engine_selector_div = selector_divs[2]
                                                select_element_engine = engine_selector_div.find_element(By.TAG_NAME,
                                                                                                         'select')
                                                options_engine = select_element_engine.find_elements(By.TAG_NAME,
                                                                                                     'option')
                                            except NoSuchElementException as e:
                                                print(f"StaleElementReferenceException: {e}")
                                                driver.refresh()
                                                sleep(5)
                                                main_div = driver.find_element(By.XPATH,
                                                                               './/div[contains(@class, "header-select__choosse-wrap")]')
                                                selector_divs = main_div.find_elements(By.XPATH,
                                                                                       './/div[contains(@class, "selector")]')
                                                engine_selector_div = selector_divs[2]
                                                select_element_engine = engine_selector_div.find_element(By.TAG_NAME,
                                                                                                         'select')
                                                options_engine = select_element_engine.find_elements(By.TAG_NAME,
                                                                                                     'option')
                                except StaleElementReferenceException:
                                    print("StaleElementReferenceException. Next car...")
                                    main_div = driver.find_element(By.XPATH,
                                                                   './/div[contains(@class, "header-select__choosse-wrap")]')
                                    selector_divs = main_div.find_elements(By.XPATH,
                                                                           './/div[contains(@class, "selector")]')
                                    engine_selector_div = selector_divs[2]
                                    select_element_engine = engine_selector_div.find_element(By.TAG_NAME, 'select')
                                    options_engine = select_element_engine.find_elements(By.TAG_NAME, 'option')
                                    continue
                        try:
                            # Ponovno pronalaženje serije nakon iteracije kroz motore
                            main_div = driver.find_element(By.XPATH,
                                                           './/div[contains(@class, "header-select__choosse-wrap")]')
                            selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                            second_selector_div = selector_divs[1]
                            select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
                            models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
                            if j < len(models):
                                model = models[j]
                                options_series = model.find_elements(By.TAG_NAME, 'option')
                                print(f"PASSED: 'j' ({j}) is in the range for models list (length {len(models)})")
                            else:
                                print(
                                    f"[800] IndexError: 'j' ({j}) is out of range for models list (length {len(models)})")
                                screenshot_path1 = "/home/nikola/Projects/Local Projects/online-car-parts-fuel-filters/error1.png"
                                driver.save_screenshot(screenshot_path1)
                        except IndexError as e:
                            try:
                                print(f"IndexError: {e}")
                                driver.refresh()
                                sleep(5)
                                main_div = driver.find_element(By.XPATH,
                                                               './/div[contains(@class, "header-select__choosse-wrap")]')
                                selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                                second_selector_div = selector_divs[1]
                                select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME,
                                                                                                   'select')
                                models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
                                if j < len(models):
                                    model = models[j]
                                    options_series = model.find_elements(By.TAG_NAME, 'option')
                                else:
                                    print(
                                        f"[820] IndexError: 'j' ({j}) is out of range for models list (length {len(models)})")
                            except IndexError as e:
                                print(f"IndexError: {e}")
                                driver.refresh()
                                sleep(5)
                                continue
                        except StaleElementReferenceException as e:
                            print(f"StaleElementReferenceException: {e}")
                            # Ponovo pronađite element i nastavite
                            driver.refresh()
                            sleep(5)
                            main_div = driver.find_element(By.XPATH,
                                                           './/div[contains(@class, "header-select__choosse-wrap")]')
                            selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                            second_selector_div = selector_divs[1]
                            select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
                            models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
                            if j < len(models):
                                model = models[j]
                                options_series = model.find_elements(By.TAG_NAME, 'option')
                            else:
                                print(
                                    f"[841] IndexError: 'j' ({j}) is out of range for models list (length {len(models)})")
                except StaleElementReferenceException as e:
                    print(f"StaleElementReferenceException: {e}")
                    driver.refresh()
                    sleep(5)
                    main_div = driver.find_element(By.XPATH,
                                                   './/div[contains(@class, "header-select__choosse-wrap")]')
                    selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                    second_selector_div = selector_divs[1]
                    select_element_model_and_series = second_selector_div.find_element(By.TAG_NAME, 'select')
                    models = select_element_model_and_series.find_elements(By.XPATH, './/optgroup')
                    if j < len(models):
                        model = models[j]
                        options_series = model.find_elements(By.TAG_NAME, 'option')
                    else:
                        print(f"[856] IndexError: 'j' ({j}) is out of range for models list (length {len(models)})")
                        screenshot_path2 = "/home/nikola/Projects/Local Projects/online-car-parts-fuel-filters/error2.png"
                        driver.save_screenshot(screenshot_path2)
            print("---------------")
            workbook.save(file_path)
            print("Brand saved")
            time_now = datetime.now()
            formatted_time = time_now.strftime("%Y-%m-%d %H:%M:%S")
            print("----------------Time now:", formatted_time)
            sleep(0.1)

            # Provjera da li ima još brendova koji nisu obrađeni
            if i < len(options_brand) - 1:

                try:
                    # Ponovno pronalaženje opcija brenda
                    alphabetical_optgroup_brand = select_element_brand.find_element(By.XPATH,
                                                                                    './/optgroup[@label="Carmakers are arranged in alphabetical order"]')
                except StaleElementReferenceException as e:
                    main_div = driver.find_element(By.XPATH, './/div[contains(@class, "header-select__choosse-wrap")]')
                    selector_divs = main_div.find_elements(By.XPATH, './/div[contains(@class, "selector")]')
                    first_selector_div = selector_divs[0]
                    select_element_brand = first_selector_div.find_element(By.TAG_NAME, 'select')
                    alphabetical_optgroup_brand = select_element_brand.find_element(By.XPATH,
                                                                                    './/optgroup[@label="Carmakers are arranged in alphabetical order"]')


# Putanja do Excel fajla
file_path = "/home/nikola/Projects/Local Projects/online-car-parts-fuel-filters/car_parts_data.xlsx"

# Inicijalizacija Excel fajla
initialize_excel(file_path)

workbook = load_workbook(file_path)
sheet = workbook.active

driver = setup_driver()
online_car_parts(driver, file_path)

# Podešavanje širine kolona
adjust_column_widths(file_path)
close_excel(file_path)
